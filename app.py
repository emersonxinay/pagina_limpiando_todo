from flask import Flask, render_template, url_for
import openpyxl
app = Flask(__name__)
# rutas 
@app.route('/')
def inicio():
      try:
          # Cargar el archivo Excel
          workbook = openpyxl.load_workbook('./static/precios1.xlsx', data_only=True)

          # Leer datos de las tres hojas y almacenarlos en una lista de listas
          datos = []
          for sheet_name in workbook.sheetnames:
              hoja = workbook[sheet_name]
              hoja_datos = []
              for row in hoja.iter_rows(values_only=True):
                  hoja_datos.append(row)
              datos.append({'hoja': sheet_name, 'datos': hoja_datos})

          # Renderizar la plantilla con los datos
          return render_template('base.html', datos=datos)
      except Exception as e:
          return f"Error: {e}"

@app.route('/datos')
def mostrar_datos():
    try:
        # Cargar el archivo Excel
        workbook = openpyxl.load_workbook('./static/resultado1.xlsx', data_only=True)

        # Leer datos de las tres hojas y almacenarlos en una lista de listas
        datos = []
        for sheet_name in workbook.sheetnames:
            hoja = workbook[sheet_name]
            hoja_datos = []
            for row in hoja.iter_rows(values_only=True):
                hoja_datos.append(row)
            datos.append({'hoja': sheet_name, 'datos': hoja_datos})

        # Renderizar la plantilla con los datos
        return render_template('base.html', datos=datos)
    except Exception as e:
        return f"Error: {e}"

# bloque de prueba 
if __name__ == "__main__":
  app.run(debug=True)